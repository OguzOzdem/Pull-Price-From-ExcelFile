import openpyxl
from CalculPrice import Calculate

class Product:

    def GetCategories(self):
        file = openpyxl.load_workbook("./sorular-excel.xlsx")
        page = file["Veriler"]
        categories = []
        column_num = page.max_column
        for column in range(1,column_num+1):
            if page.cell(1, column).value == None:
                continue

            categories.append(page.cell(1, column).value)

        return categories


    def FindProduct(self,claim):
        file = openpyxl.load_workbook("./sorular-excel.xlsx")
        page = file["Veriler"]

        row_num = page.max_row
        column_num= page.max_column

        claim = claim.upper()

        CC = Calculate
        Currency = CC.CalculateCurrency(self)

        list = []

        for row in range(2,row_num+1):
            product_column = 2
            price_column = 6

            data = page.cell(row,product_column).value
            data = data.upper()
            word = ""

            for i in range(len(data)):
                word = word + data[i]
                if claim == word:



                    price = page.cell(row,price_column).value
                    price_tl = f"{price:,.3f}"
                    price_dolar = f"{(Currency * price):,.3f}"

                    print(
                        "fiyat = {} tl, {} $".format(price_tl,price_dolar))

                    list.append(data)
                    list.append(price_tl)
                    list.append(price_dolar)
                if data[i] == " ":
                    word = ""

        return list